"""
Dashboard Service — PB159 đến PB189
"""
import io
from calendar import monthrange
from datetime import datetime, timezone, timedelta
from typing import Optional
from uuid import UUID

from fastapi import HTTPException
from sqlalchemy.orm import Session

from app.models.task import Task, TaskAssignee
from app.models.user import User
from app.models.organization import Department
from app.models.notification import Notification


# ── Helpers ───────────────────────────────────────────────────

def _make_aware(dt):
    if dt and dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt


def _get_week_range(dt: datetime):
    """Trả về đầu tuần (Mon) và cuối tuần (Sun)"""
    start = dt - timedelta(days=dt.weekday())
    end = start + timedelta(days=6)
    return start.replace(hour=0, minute=0, second=0), end.replace(hour=23, minute=59, second=59)


def _get_month_range(year: int, month: int):
    last_day = monthrange(year, month)[1]
    start = datetime(year, month, 1, tzinfo=timezone.utc)
    end = datetime(year, month, last_day, 23, 59, 59, tzinfo=timezone.utc)
    return start, end


def _prev_month(year: int, month: int):
    if month == 1:
        return year - 1, 12
    return year, month - 1


def _is_overdue(task: Task, now: datetime) -> bool:
    deadline = _make_aware(task.deadline)
    return (deadline is not None
            and deadline < now
            and task.status not in ("done", "cancelled"))


def _days_late(task: Task, now: datetime) -> int:
    deadline = _make_aware(task.deadline)
    if not deadline or deadline >= now:
        return 0
    return (now - deadline).days


def _enrich_task(task: Task, db: Session, now: datetime) -> dict:
    assignees = []
    for ta in task.assignees:
        u = db.query(User).filter(User.id == ta.user_id).first()
        if u:
            assignees.append({"user_id": str(u.id), "full_name": u.full_name, "avatar_url": u.avatar_url})
    return {
        "id": str(task.id),
        "title": task.title,
        "status": task.status,
        "priority": task.priority,
        "progress_pct": task.progress_pct,
        "deadline": task.deadline.isoformat() if task.deadline else None,
        "created_at": task.created_at.isoformat() if task.created_at else None,
        "is_overdue": _is_overdue(task, now),
        "days_late": _days_late(task, now),
        "assignees": assignees,
    }


# ── Gantt Chart ───────────────────────────────────────────────

def get_gantt(dept_id: UUID, view: str, db: Session) -> dict:
    """PB159, PB160"""
    now = datetime.now(timezone.utc)

    if view == "day":
        start = now.replace(hour=0, minute=0, second=0)
        end = now.replace(hour=23, minute=59, second=59)
    elif view == "week":
        start, end = _get_week_range(now)
    else:  # month
        start, end = _get_month_range(now.year, now.month)

    tasks = db.query(Task).filter(
        Task.dept_id == dept_id,
        Task.status != "cancelled",
    ).all()

    return {
        "view": view,
        "start": start.isoformat(),
        "end": end.isoformat(),
        "tasks": [_enrich_task(t, db, now) for t in tasks],
    }


# ── Calendar View ─────────────────────────────────────────────

def get_calendar_month(user: User, year: int, month: int, db: Session) -> dict:
    """PB162"""
    now = datetime.now(timezone.utc)
    start, end = _get_month_range(year, month)
    last_day = monthrange(year, month)[1]

    # Lấy task theo role
    if user.role == "staff":
        assigned_ids = [ta.task_id for ta in db.query(TaskAssignee).filter(
            TaskAssignee.user_id == user.id
        ).all()]
        tasks = db.query(Task).filter(
            Task.id.in_(assigned_ids),
            Task.deadline >= start,
            Task.deadline <= end,
        ).all()
    else:
        tasks = db.query(Task).filter(
            Task.dept_id == user.dept_id,
            Task.deadline >= start,
            Task.deadline <= end,
        ).all()

    # Group theo ngày
    days = []
    for day_num in range(1, last_day + 1):
        day_tasks = [
            _enrich_task(t, db, now) for t in tasks
            if t.deadline and _make_aware(t.deadline).day == day_num
        ]
        days.append({
            "day": day_num,
            "date": f"{year}-{month:02d}-{day_num:02d}",
            "tasks": day_tasks,
            "task_count": len(day_tasks),
        })

    return {"year": year, "month": month, "days": days}


def get_calendar_week(user: User, date_str: str, db: Session) -> dict:
    """PB163"""
    now = datetime.now(timezone.utc)
    date = datetime.fromisoformat(date_str).replace(tzinfo=timezone.utc)
    week_start, week_end = _get_week_range(date)

    if user.role == "staff":
        assigned_ids = [ta.task_id for ta in db.query(TaskAssignee).filter(
            TaskAssignee.user_id == user.id
        ).all()]
        tasks = db.query(Task).filter(
            Task.id.in_(assigned_ids),
            Task.deadline >= week_start,
            Task.deadline <= week_end,
        ).all()
    else:
        tasks = db.query(Task).filter(
            Task.dept_id == user.dept_id,
            Task.deadline >= week_start,
            Task.deadline <= week_end,
        ).all()

    days = []
    for i in range(7):
        day = week_start + timedelta(days=i)
        day_tasks = [
            _enrich_task(t, db, now) for t in tasks
            if t.deadline and _make_aware(t.deadline).date() == day.date()
        ]
        days.append({
            "date": day.date().isoformat(),
            "weekday": day.strftime("%A"),
            "tasks": day_tasks,
        })

    return {"week_start": week_start.date().isoformat(), "days": days}


def get_calendar_day(user: User, date_str: str, db: Session) -> dict:
    """PB164"""
    now = datetime.now(timezone.utc)
    date = datetime.fromisoformat(date_str).replace(tzinfo=timezone.utc)
    day_start = date.replace(hour=0, minute=0, second=0)
    day_end = date.replace(hour=23, minute=59, second=59)

    if user.role == "staff":
        assigned_ids = [ta.task_id for ta in db.query(TaskAssignee).filter(
            TaskAssignee.user_id == user.id
        ).all()]
        tasks = db.query(Task).filter(
            Task.id.in_(assigned_ids),
            Task.deadline >= day_start,
            Task.deadline <= day_end,
        ).all()
    else:
        tasks = db.query(Task).filter(
            Task.dept_id == user.dept_id,
            Task.deadline >= day_start,
            Task.deadline <= day_end,
        ).all()

    return {
        "date": date_str,
        "tasks": [_enrich_task(t, db, now) for t in tasks],
    }


# ── Performance Report ────────────────────────────────────────

def get_performance_report(user: User, year: int, month: int,
                            from_date: Optional[datetime], to_date: Optional[datetime],
                            db: Session) -> dict:
    """PB165, PB166, PB167, PB170, PB171"""
    now = datetime.now(timezone.utc)

    if from_date and to_date:
        start = _make_aware(from_date)
        end = _make_aware(to_date).replace(hour=23, minute=59, second=59)
    else:
        start, end = _get_month_range(year, month)

    staff_list = db.query(User).filter(
        User.dept_id == user.dept_id,
        User.role == "staff",
        User.is_active == True,
    ).all()

    staff_perf = []
    for s in staff_list:
        assigned_ids = [ta.task_id for ta in db.query(TaskAssignee).filter(
            TaskAssignee.user_id == s.id
        ).all()]

        tasks_in_period = db.query(Task).filter(
            Task.id.in_(assigned_ids),
            Task.deadline >= start,
            Task.deadline <= end,
            Task.status != "cancelled",
        ).all()

        done_tasks = [t for t in tasks_in_period if t.status == "done"]
        on_time = sum(
            1 for t in done_tasks
            if t.completed_at and t.deadline and
            _make_aware(t.completed_at) <= _make_aware(t.deadline)
        )
        on_time_rate = round(on_time / len(done_tasks) * 100, 2) if done_tasks else 0.0

        # PB167: thời gian trung bình hoàn thành (ngày)
        completion_times = []
        for t in done_tasks:
            if t.completed_at and t.created_at:
                days = (_make_aware(t.completed_at) - _make_aware(t.created_at)).days
                completion_times.append(days)
        avg_days = round(sum(completion_times) / len(completion_times), 1) if completion_times else 0.0

        staff_perf.append({
            "user_id": str(s.id),
            "full_name": s.full_name,
            "avatar_url": s.avatar_url,
            "tasks_total": len(tasks_in_period),
            "tasks_done": len(done_tasks),
            "on_time_rate": on_time_rate,
            "avg_completion_days": avg_days,
        })

    return {
        "year": year,
        "month": month,
        "from_date": start.isoformat(),
        "to_date": end.isoformat(),
        "staff_performance": staff_perf,
    }


def get_overdue_by_dept(org_id: UUID, year: int, month: int, db: Session) -> list:
    """PB168"""
    now = datetime.now(timezone.utc)
    start, end = _get_month_range(year, month)

    depts = db.query(Department).filter(
        Department.org_id == org_id,
        Department.is_active == True,
    ).all()

    result = []
    for dept in depts:
        all_tasks = db.query(Task).filter(
            Task.dept_id == dept.id,
            Task.deadline >= start,
            Task.deadline <= end,
            Task.status != "cancelled",
        ).all()
        overdue = [t for t in all_tasks if _is_overdue(t, now)]
        overdue_rate = round(len(overdue) / len(all_tasks) * 100, 2) if all_tasks else 0.0

        result.append({
            "dept_id": str(dept.id),
            "dept_name": dept.name,
            "total_tasks": len(all_tasks),
            "overdue_count": len(overdue),
            "overdue_rate": overdue_rate,
        })

    return sorted(result, key=lambda x: x["overdue_rate"], reverse=True)


def get_kpi_comparison_by_quarter(org_id: UUID, year: int, quarter: int, db: Session) -> dict:
    """PB169"""
    from app.models.kpi import KpiScore

    start_month = (quarter - 1) * 3 + 1
    months = [start_month, start_month + 1, start_month + 2]

    depts = db.query(Department).filter(
        Department.org_id == org_id,
        Department.is_active == True,
    ).all()

    dept_data = []
    for dept in depts:
        members = db.query(User).filter(
            User.dept_id == dept.id,
            User.role == "staff",
            User.is_active == True,
        ).all()

        month_scores = []
        for m in months:
            scores = []
            for member in members:
                kpi_scores = db.query(KpiScore).filter(
                    KpiScore.user_id == member.id,
                    KpiScore.year == year,
                    KpiScore.month == m,
                ).all()
                total = sum(s.weighted_score for s in kpi_scores)
                scores.append(total)
            avg = round(sum(scores) / len(scores), 2) if scores else 0.0
            month_scores.append({"month": m, "avg_score": avg})

        quarter_avg = round(
            sum(ms["avg_score"] for ms in month_scores) / len(month_scores), 2
        ) if month_scores else 0.0

        dept_data.append({
            "dept_id": str(dept.id),
            "dept_name": dept.name,
            "avg_score": quarter_avg,
            "months": month_scores,
        })

    return {
        "year": year,
        "quarter": quarter,
        "departments": sorted(dept_data, key=lambda x: x["avg_score"], reverse=True),
    }


# ── Export ────────────────────────────────────────────────────

def export_performance_excel(user: User, year: int, month: int, db: Session) -> bytes:
    """PB172"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    report = get_performance_report(user, year, month, None, None, db)

    wb = openpyxl.Workbook()
    ws = wb.active
    safe_title = f"Hiệu suất {month}-{year}"
    for ch in ['\\', '/', '*', '?', ':', '[', ']']:
        safe_title = safe_title.replace(ch, '-')
    ws.title = safe_title[:31]

    headers = ["Họ tên", "Tổng task", "Task hoàn thành",
               "Tỉ lệ đúng hạn (%)", "Thời gian TB (ngày)"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 20

    for s in report["staff_performance"]:
        ws.append([
            s["full_name"],
            s["tasks_total"],
            s["tasks_done"],
            s["on_time_rate"],
            s["avg_completion_days"],
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def export_performance_pdf(user: User, year: int, month: int, db: Session) -> bytes:
    """PB173: tạo PDF đơn giản bằng reportlab hoặc HTML → PDF"""
    report = get_performance_report(user, year, month, None, None, db)

    # Tạo HTML rồi convert sang PDF đơn giản
    html = f"""
    <html><head><meta charset="utf-8">
    <style>
      body {{ font-family: Arial, sans-serif; padding: 20px; }}
      h1 {{ color: #4F46E5; }} table {{ width: 100%; border-collapse: collapse; }}
      th {{ background: #4F46E5; color: white; padding: 8px; }}
      td {{ padding: 6px; border: 1px solid #ddd; }}
      tr:nth-child(even) {{ background: #f9f9f9; }}
    </style></head><body>
    <h1>Báo cáo hiệu suất tháng {month}/{year}</h1>
    <p>Phòng ban | Số nhân viên: {len(report["staff_performance"])}</p>
    <table>
      <tr><th>Họ tên</th><th>Task hoàn thành</th>
          <th>Tỉ lệ đúng hạn</th><th>Thời gian TB</th></tr>
    """
    for s in report["staff_performance"]:
        html += f"""
      <tr><td>{s['full_name']}</td><td>{s['tasks_done']}/{s['tasks_total']}</td>
          <td>{s['on_time_rate']}%</td><td>{s['avg_completion_days']} ngày</td></tr>"""
    html += "</table></body></html>"

    try:
        import weasyprint
        pdf_bytes = weasyprint.HTML(string=html).write_pdf()
    except ImportError:
        # Fallback: trả về HTML bytes với content-type pdf
        pdf_bytes = html.encode("utf-8")

    return pdf_bytes


# ── CEO Dashboard ─────────────────────────────────────────────

def get_ceo_dashboard(org_id: UUID, db: Session) -> dict:
    """PB174-PB178, PB180"""
    now = datetime.now(timezone.utc)
    from app.models.kpi import KpiScore

        # PB174: tổng nhân sự active trong org, không tính CEO
    total_employees = db.query(User).filter(
        User.org_id == org_id,
            User.role != "ceo",
        User.is_active == True,
    ).count()

    # PB175: task theo trạng thái
    all_tasks = db.query(Task).join(
        Department, Task.dept_id == Department.id
    ).filter(
        Department.org_id == org_id,
        Task.status != "cancelled",
    ).all()

    task_stats = {
        "todo": sum(1 for t in all_tasks if t.status == "todo"),
        "in_progress": sum(1 for t in all_tasks if t.status == "in_progress"),
        "done": sum(1 for t in all_tasks if t.status == "done"),
        "total": len(all_tasks),
    }

    # PB176: task overdue
    overdue_count = sum(1 for t in all_tasks if _is_overdue(t, now))

    # PB177: KPI trung bình theo phòng ban
    depts = db.query(Department).filter(
        Department.org_id == org_id,
        Department.is_active == True,
    ).all()

    kpi_by_dept = []
    for dept in depts:
        members = db.query(User).filter(
            User.dept_id == dept.id,
            User.role == "staff",
            User.is_active == True,
        ).all()
        scores = []
        for m in members:
            kpi_scores = db.query(KpiScore).filter(
                KpiScore.user_id == m.id,
                KpiScore.year == now.year,
                KpiScore.month == now.month,
            ).all()
            scores.append(sum(s.weighted_score for s in kpi_scores))
        avg = round(sum(scores) / len(scores), 2) if scores else 0.0
        kpi_by_dept.append({
            "dept_id": str(dept.id),
            "dept_name": dept.name,
            "avg_kpi_score": avg,
            "member_count": len(members),
        })
    kpi_by_dept.sort(key=lambda x: x["avg_kpi_score"], reverse=True)

    # PB178: top 30 nhân viên KPI cao nhất
    all_staff = db.query(User).filter(
        User.org_id == org_id,
        User.role == "staff",
        User.is_active == True,
    ).all()

    staff_scores = []
    for s in all_staff:
        kpi_scores = db.query(KpiScore).filter(
            KpiScore.user_id == s.id,
            KpiScore.year == now.year,
            KpiScore.month == now.month,
        ).all()
        total = sum(sc.weighted_score for sc in kpi_scores)
        dept = db.query(Department).filter(Department.id == s.dept_id).first()
        staff_scores.append({
            "user_id": str(s.id),
            "full_name": s.full_name,
            "avatar_url": s.avatar_url,
            "dept_name": dept.name if dept else "",
            "kpi_score": round(total, 2),
        })
    staff_scores.sort(key=lambda x: x["kpi_score"], reverse=True)
    top_employees = staff_scores[:30]

    return {
        "total_employees": total_employees,
        "task_stats": task_stats,
        "overdue_tasks": overdue_count,
        "kpi_by_dept": kpi_by_dept,
        "top_employees": top_employees,
    }


def get_ceo_heatmap(org_id: UUID, year: int, month: int, db: Session) -> dict:
    """PB179"""
    start, end = _get_month_range(year, month)
    last_day = monthrange(year, month)[1]

    done_tasks = db.query(Task).join(
        Department, Task.dept_id == Department.id
    ).filter(
        Department.org_id == org_id,
        Task.status == "done",
        Task.completed_at >= start,
        Task.completed_at <= end,
    ).all()

    heatmap = []
    for day_num in range(1, last_day + 1):
        count = sum(
            1 for t in done_tasks
            if t.completed_at and _make_aware(t.completed_at).day == day_num
        )
        heatmap.append({
            "date": f"{year}-{month:02d}-{day_num:02d}",
            "day": day_num,
            "tasks_done": count,
        })

    return {"year": year, "month": month, "heatmap": heatmap}


def get_system_usage(org_id: UUID, db: Session) -> dict:
    """PB180"""
    from app.models.user import LoginLog
    now = datetime.now(timezone.utc)

    def count_active(days: int) -> int:
        since = now - timedelta(days=days)
        return db.query(LoginLog.user_id).filter(
            LoginLog.success == True,
            LoginLog.created_at >= since,
        ).distinct().count()

    return {
        "daily_active_users": count_active(1),
        "weekly_active_users": count_active(7),
        "monthly_active_users": count_active(30),
    }


# ── Manager Dashboard ─────────────────────────────────────────

def get_manager_dashboard(user: User, db: Session) -> dict:
    """PB181-PB186"""
    now = datetime.now(timezone.utc)
    week_start, week_end = _get_week_range(now)
    month_start, month_end = _get_month_range(now.year, now.month)

    dept_tasks = db.query(Task).filter(
        Task.dept_id == user.dept_id,
        Task.status != "cancelled",
    ).all()

    # PB181: task stats
    task_stats = {
        "todo": sum(1 for t in dept_tasks if t.status == "todo"),
        "in_progress": sum(1 for t in dept_tasks if t.status == "in_progress"),
        "done": sum(1 for t in dept_tasks if t.status == "done"),
        "total": len(dept_tasks),
    }

    # PB182: overdue
    overdue_tasks = [t for t in dept_tasks if _is_overdue(t, now)]
    overdue_list = [_enrich_task(t, db, now) for t in overdue_tasks]

    # PB183: workload per staff
    staff_members = db.query(User).filter(
        User.dept_id == user.dept_id,
        User.role == "staff",
        User.is_active == True,
    ).all()

    workload = []
    for s in staff_members:
        assigned = db.query(TaskAssignee).filter(TaskAssignee.user_id == s.id).count()
        workload.append({
            "user_id": str(s.id),
            "full_name": s.full_name,
            "avatar_url": s.avatar_url,
            "task_count": assigned,
        })
    workload.sort(key=lambda x: x["task_count"], reverse=True)

    # PB184: weekly progress
    week_tasks = db.query(Task).filter(
        Task.dept_id == user.dept_id,
        Task.deadline >= week_start,
        Task.deadline <= week_end,
        Task.status != "cancelled",
    ).all()
    done_week = sum(1 for t in week_tasks if t.status == "done")
    completion_rate = round(done_week / len(week_tasks) * 100, 1) if week_tasks else 0.0

    weekly_progress = {
        "done_this_week": done_week,
        "total_this_week": len(week_tasks),
        "completion_rate": completion_rate,
    }

    # PB185: top overdue tasks
    top_overdue = sorted(overdue_tasks, key=lambda t: _days_late(t, now), reverse=True)[:5]
    top_overdue_list = []
    for t in top_overdue:
        enriched = _enrich_task(t, db, now)
        enriched["days_late"] = _days_late(t, now)
        top_overdue_list.append(enriched)

    # PB186: so sánh tháng này vs tháng trước
    prev_year, prev_month = _prev_month(now.year, now.month)
    prev_start, prev_end = _get_month_range(prev_year, prev_month)

    curr_done = sum(
        1 for t in dept_tasks
        if t.status == "done" and t.completed_at and
        month_start <= _make_aware(t.completed_at) <= month_end
    )
    prev_tasks = db.query(Task).filter(
        Task.dept_id == user.dept_id,
        Task.status == "done",
        Task.completed_at >= prev_start,
        Task.completed_at <= prev_end,
    ).all()
    prev_done = len(prev_tasks)

    change = curr_done - prev_done
    direction = "up" if change > 0 else ("down" if change < 0 else "same")

    return {
        "task_stats": task_stats,
        "overdue_tasks": len(overdue_tasks),
        "overdue_task_list": overdue_list,
        "workload": workload,
        "weekly_progress": weekly_progress,
        "top_overdue_tasks": top_overdue_list,
        "month_comparison": {
            "tasks_done_this_month": curr_done,
            "tasks_done_last_month": prev_done,
            "tasks_done_change": change,
            "on_time_rate_change": 0.0,  # placeholder
            "direction": direction,
        },
    }


# ── Staff Dashboard ───────────────────────────────────────────

def get_staff_dashboard(user: User, db: Session) -> dict:
    """PB187-PB189"""
    now = datetime.now(timezone.utc)
    today_start = now.replace(hour=0, minute=0, second=0)
    today_end = now.replace(hour=23, minute=59, second=59)
    month_start, month_end = _get_month_range(now.year, now.month)
    prev_year, prev_month = _prev_month(now.year, now.month)
    prev_start, prev_end = _get_month_range(prev_year, prev_month)

    assigned_ids = [ta.task_id for ta in db.query(TaskAssignee).filter(
        TaskAssignee.user_id == user.id
    ).all()]

    all_tasks = db.query(Task).filter(
        Task.id.in_(assigned_ids),
        Task.status != "cancelled",
    ).all()

    # PB187: task hôm nay (deadline hôm nay hoặc in_progress)
    tasks_today = [
        t for t in all_tasks
        if (t.deadline and today_start <= _make_aware(t.deadline) <= today_end)
        or t.status == "in_progress"
    ]
    tasks_today.sort(key=lambda t: _make_aware(t.deadline) if t.deadline else datetime.max.replace(tzinfo=timezone.utc))
    tasks_today_data = [_enrich_task(t, db, now) for t in tasks_today]

    # PB189: task done tháng này
    done_this_month = db.query(Task).filter(
        Task.id.in_(assigned_ids),
        Task.status == "done",
        Task.completed_at >= month_start,
        Task.completed_at <= month_end,
    ).count()

    done_last_month = db.query(Task).filter(
        Task.id.in_(assigned_ids),
        Task.status == "done",
        Task.completed_at >= prev_start,
        Task.completed_at <= prev_end,
    ).count()

    change = done_this_month - done_last_month
    direction = "up" if change > 0 else ("down" if change < 0 else "same")

    # PB188: KPI tháng hiện tại
    try:
        from app.services.kpi_service import get_my_kpi
        kpi_data = get_my_kpi(user, now.year, now.month, db)
        kpi_current = {
            "total_score": kpi_data["total_score"],
            "target_score": kpi_data["target_score"],
            "grade": kpi_data["grade"],
        }
    except Exception:
        kpi_current = {"total_score": 0.0, "target_score": 75.0, "grade": "Chưa có dữ liệu"}

    return {
        "tasks_today": tasks_today_data,
        "tasks_done_this_month": done_this_month,
        "tasks_done_last_month": done_last_month,
        "tasks_done_change": change,
        "change_direction": direction,
        "kpi_current_month": kpi_current,
    }
