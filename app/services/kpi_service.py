"""
KPI Service — PB120 đến PB158
"""
import io
from datetime import datetime, timezone, timedelta
from typing import Optional
from uuid import UUID

from fastapi import HTTPException
from sqlalchemy.orm import Session

from app.models.kpi import (
    KpiConfig, KpiCriteria, KpiCriteriaHistory,
    KpiScore, KpiTarget, KpiFinalize,
    KpiAppeal, KpiAdjustment,
)
from app.models.notification import Notification
from app.models.organization import Department
from app.models.task import Task, TaskAssignee
from app.models.user import User


# ── Helpers ───────────────────────────────────────────────────

def _get_config(org_id: UUID, db: Session) -> KpiConfig:
    cfg = db.query(KpiConfig).filter(KpiConfig.org_id == org_id).first()
    if not cfg:
        # Tạo config mặc định nếu chưa có
        cfg = KpiConfig(
            org_id=org_id,
            target_score=75.0, cycle_day=1,
            threshold_excellent=90.0, threshold_good=75.0, threshold_pass=60.0,
        )
        db.add(cfg)
        db.commit()
        db.refresh(cfg)
    return cfg


def _grade(score: float, cfg: KpiConfig) -> str:
    if score >= cfg.threshold_excellent:
        return "Xuất sắc"
    elif score >= cfg.threshold_good:
        return "Tốt"
    elif score >= cfg.threshold_pass:
        return "Đạt"
    else:
        return "Chưa đạt"


def _is_finalized(org_id: UUID, year: int, month: int, db: Session) -> bool:
    rec = db.query(KpiFinalize).filter(
        KpiFinalize.org_id == org_id,
        KpiFinalize.year == year,
        KpiFinalize.month == month,
        KpiFinalize.is_finalized == True,
    ).first()
    return rec is not None


def _assert_not_finalized(org_id: UUID, year: int, month: int, db: Session):
    if _is_finalized(org_id, year, month, db):
        raise HTTPException(status_code=403, detail="KPI tháng này đã được chốt, không thể chỉnh sửa")


# ── KPI Config ────────────────────────────────────────────────

def save_config(org_id: UUID, target_score: float, cycle_day: int,
                thresholds: dict, user_id: UUID, db: Session) -> dict:
    """PB122, PB123, PB124"""
    cfg = db.query(KpiConfig).filter(KpiConfig.org_id == org_id).first()
    if not cfg:
        cfg = KpiConfig(org_id=org_id)
        db.add(cfg)

    cfg.target_score = target_score
    cfg.cycle_day = cycle_day
    cfg.threshold_excellent = thresholds.get("excellent", 90)
    cfg.threshold_good = thresholds.get("good", 75)
    cfg.threshold_pass = thresholds.get("pass", 60)
    cfg.updated_by = user_id
    db.commit()
    db.refresh(cfg)

    return {
        "target_score": cfg.target_score,
        "cycle_day": cfg.cycle_day,
        "thresholds": {
            "excellent": cfg.threshold_excellent,
            "good": cfg.threshold_good,
            "pass": cfg.threshold_pass,
        }
    }


def get_config(org_id: UUID, db: Session) -> dict:
    cfg = _get_config(org_id, db)
    return {
        "target_score": cfg.target_score,
        "cycle_day": cfg.cycle_day,
        "thresholds": {
            "excellent": cfg.threshold_excellent,
            "good": cfg.threshold_good,
            "pass": cfg.threshold_pass,
        }
    }


# ── KPI Criteria ──────────────────────────────────────────────

def create_criteria(org_id: UUID, dept_id: Optional[UUID], name: str, description: Optional[str],
                    weight: float, is_global: bool, formula_type: str,
                    user: User, db: Session) -> KpiCriteria:
    """PB120, PB125"""
    c = KpiCriteria(
        org_id=org_id, dept_id=dept_id,
        name=name, description=description,
        weight=weight, default_weight=weight,
        is_global=is_global, formula_type=formula_type,
        created_by=user.id,
    )
    db.add(c)
    db.commit()
    db.refresh(c)
    return c


def update_criteria(criteria_id: UUID, weight: Optional[float], name: Optional[str],
                    description: Optional[str], user: User, db: Session) -> KpiCriteria:
    """PB126: điều chỉnh trọng số trong biên độ ±20 điểm (absolute)"""
    c = db.query(KpiCriteria).filter(KpiCriteria.id == criteria_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Không tìm thấy tiêu chí KPI")

    if weight is not None:
        # PB126: Manager chỉ điều chỉnh trong biên độ ±20 điểm so với default (absolute)
        if user.role == "manager" and c.default_weight is not None:
            max_allowed = min(100.0, c.default_weight + 20.0)
            min_allowed = max(0.0, c.default_weight - 20.0)
            if not (min_allowed <= weight <= max_allowed):
                raise HTTPException(
                    status_code=400,
                    detail=f"Manager chỉ điều chỉnh trong biên độ ±20 điểm (từ {min_allowed:.1f}% đến {max_allowed:.1f}%)"
                )

        # PB128: ghi log thay đổi
        if c.weight != weight:
            db.add(KpiCriteriaHistory(
                criteria_id=c.id, changed_by=user.id,
                old_weight=c.weight, new_weight=weight,
                old_name=c.name, new_name=name or c.name,
            ))
        c.weight = weight

    if name:
        c.name = name
    if description is not None:
        c.description = description

    db.commit()
    db.refresh(c)
    return c


def list_criteria(org_id: UUID, dept_id: Optional[UUID], db: Session) -> list:
    q = db.query(KpiCriteria).filter(
        KpiCriteria.org_id == org_id,
        KpiCriteria.is_active == True,
    )
    if dept_id:
        q = q.filter(
            (KpiCriteria.is_global == True) |
            (KpiCriteria.dept_id == dept_id)
        )
    else:
        q = q.filter(KpiCriteria.is_global == True)
    return q.all()


def validate_weights(org_id: UUID, dept_id: Optional[UUID], db: Session) -> dict:
    """PB121, PB127"""
    criteria = list_criteria(org_id, dept_id, db)
    total = sum(c.weight for c in criteria)
    is_valid = abs(total - 100.0) < 0.01

    result = {"total_weight": round(total, 2), "is_valid": is_valid}
    if not is_valid:
        result["warning"] = f"Tổng trọng số hiện tại là {total:.1f}%, phải bằng 100%"
    return result


def get_criteria_history(criteria_id: UUID, db: Session) -> list:
    """PB128"""
    history = db.query(KpiCriteriaHistory).filter(
        KpiCriteriaHistory.criteria_id == criteria_id
    ).order_by(KpiCriteriaHistory.changed_at.desc()).all()

    result = []
    for h in history:
        changer = db.query(User).filter(User.id == h.changed_by).first()
        result.append({
            "id": h.id,
            "changed_by": h.changed_by,
            "changer_name": changer.full_name if changer else "",
            "old_weight": h.old_weight,
            "new_weight": h.new_weight,
            "changed_at": h.changed_at,
        })
    return result


# ── KPI Calculation Formulas ──────────────────────────────────

def calculate_on_time_rate(user_id: UUID, year: int, month: int, db: Session) -> float:
    """PB130: tỉ lệ hoàn thành task đúng hạn trong tháng"""
    from calendar import monthrange
    first_day = datetime(year, month, 1, tzinfo=timezone.utc)
    last_day_num = monthrange(year, month)[1]
    last_day = datetime(year, month, last_day_num, 23, 59, 59, tzinfo=timezone.utc)

    assigned_ids = [
        ta.task_id for ta in db.query(TaskAssignee).filter(
            TaskAssignee.user_id == user_id
        ).all()
    ]

    # Task done có deadline trong tháng
    tasks_with_deadline = db.query(Task).filter(
        Task.id.in_(assigned_ids),
        Task.status == "done",
        Task.deadline >= first_day,
        Task.deadline <= last_day,
        Task.completed_at.isnot(None),
    ).all()

    if not tasks_with_deadline:
        return 0.0

    on_time = sum(
        1 for t in tasks_with_deadline
        if t.completed_at and t.deadline and
        (t.completed_at.replace(tzinfo=timezone.utc) if t.completed_at.tzinfo is None else t.completed_at) <=
        (t.deadline.replace(tzinfo=timezone.utc) if t.deadline.tzinfo is None else t.deadline)
    )

    return round(on_time / len(tasks_with_deadline) * 100, 2)


def calculate_completion_rate(user_id: UUID, year: int, month: int, db: Session) -> float:
    """PB131: tỉ lệ hoàn thành task có deadline trong tháng"""
    from calendar import monthrange
    first_day = datetime(year, month, 1, tzinfo=timezone.utc)
    last_day_num = monthrange(year, month)[1]
    last_day = datetime(year, month, last_day_num, 23, 59, 59, tzinfo=timezone.utc)

    assigned_ids = [
        ta.task_id for ta in db.query(TaskAssignee).filter(
            TaskAssignee.user_id == user_id
        ).all()
    ]

    tasks = db.query(Task).filter(
        Task.id.in_(assigned_ids),
        Task.deadline >= first_day,
        Task.deadline <= last_day,
        Task.status != "cancelled",
    ).all()

    if not tasks:
        return 0.0

    done = sum(1 for t in tasks if t.status == "done")
    return round(done / len(tasks) * 100, 2)


def calculate_quality_rate(user_id: UUID, year: int, month: int, db: Session) -> float:
    """PB132: chất lượng dựa trên % hoàn thành trung bình"""
    from calendar import monthrange
    first_day = datetime(year, month, 1, tzinfo=timezone.utc)
    last_day_num = monthrange(year, month)[1]
    last_day = datetime(year, month, last_day_num, 23, 59, 59, tzinfo=timezone.utc)

    assigned_ids = [
        ta.task_id for ta in db.query(TaskAssignee).filter(
            TaskAssignee.user_id == user_id
        ).all()
    ]

    tasks = db.query(Task).filter(
        Task.id.in_(assigned_ids),
        Task.deadline >= first_day,
        Task.deadline <= last_day,
        Task.status != "cancelled",
    ).all()

    if not tasks:
        return 0.0

    avg = sum(t.progress_pct for t in tasks) / len(tasks)
    return round(avg, 2)


def calculate_kpi_for_user(user_id: UUID, year: int, month: int, db: Session) -> dict:
    """PB129: tính điểm KPI tổng hợp"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Không tìm thấy người dùng")

    criteria = list_criteria(user.org_id, user.dept_id, db)
    if not criteria:
        return {"total_score": 0.0, "breakdown": []}

    formula_map = {
        "on_time_rate": calculate_on_time_rate,
        "completion_rate": calculate_completion_rate,
        "quality_rate": calculate_quality_rate,
    }

    breakdown = []
    total = 0.0

    for c in criteria:
        if c.formula_type in formula_map:
            raw_score = formula_map[c.formula_type](user_id, year, month, db)
        else:
            # manual: lấy từ DB nếu có
            existing = db.query(KpiScore).filter(
                KpiScore.user_id == user_id,
                KpiScore.criteria_id == c.id,
                KpiScore.year == year,
                KpiScore.month == month,
            ).first()
            raw_score = existing.score if existing else 0.0

        weighted = raw_score * (c.weight / 100)
        total += weighted

        # Upsert KpiScore
        existing = db.query(KpiScore).filter(
            KpiScore.user_id == user_id,
            KpiScore.criteria_id == c.id,
            KpiScore.year == year,
            KpiScore.month == month,
        ).first()

        if existing and not existing.is_finalized:
            existing.score = raw_score
            existing.weighted_score = weighted
        elif not existing:
            db.add(KpiScore(
                user_id=user_id, criteria_id=c.id,
                year=year, month=month,
                score=raw_score, weighted_score=weighted,
            ))

        breakdown.append({
            "criteria_id": c.id,
            "name": c.name,
            "weight": c.weight,
            "score": raw_score,
            "weighted_score": round(weighted, 2),
            "formula_type": c.formula_type,
        })

    db.commit()
    return {"total_score": round(total, 2), "breakdown": breakdown}


# ── Staff view KPI ────────────────────────────────────────────

def get_my_kpi(user: User, year: int, month: int, db: Session) -> dict:
    """PB133-PB139"""
    result = calculate_kpi_for_user(user.id, year, month, db)
    cfg = _get_config(user.org_id, db)

    grade = _grade(result["total_score"], cfg) if result["total_score"] > 0 else "Chưa có dữ liệu"

    # Lấy target cá nhân
    target = db.query(KpiTarget).filter(
        KpiTarget.user_id == user.id,
        KpiTarget.year == year,
        KpiTarget.month == month,
    ).first()

    return {
        "user_id": user.id,
        "full_name": user.full_name,
        "year": year,
        "month": month,
        "total_score": result["total_score"],
        "grade": grade,
        "target_score": target.target_score if target else cfg.target_score,
        "breakdown": result["breakdown"],
    }


def get_kpi_history(user_id: UUID, months: int, db: Session) -> list:
    """PB136, PB144"""
    now = datetime.now(timezone.utc)
    result = []

    for i in range(months):
        month = now.month - i
        year = now.year
        if month <= 0:
            month += 12
            year -= 1

        scores = db.query(KpiScore).filter(
            KpiScore.user_id == user_id,
            KpiScore.year == year,
            KpiScore.month == month,
        ).all()

        total = sum(s.weighted_score for s in scores)
        result.append({"year": year, "month": month, "total_score": round(total, 2)})

    return result


def compare_with_dept(user: User, year: int, month: int, db: Session) -> dict:
    """PB137"""
    my_result = calculate_kpi_for_user(user.id, year, month, db)

    dept_members = db.query(User).filter(
        User.dept_id == user.dept_id,
        User.role == "staff",
        User.is_active == True,
    ).all()

    dept_scores = []
    for m in dept_members:
        scores = db.query(KpiScore).filter(
            KpiScore.user_id == m.id,
            KpiScore.year == year,
            KpiScore.month == month,
        ).all()
        dept_scores.append(sum(s.weighted_score for s in scores))

    dept_avg = sum(dept_scores) / len(dept_scores) if dept_scores else 0.0

    return {
        "my_score": my_result["total_score"],
        "dept_average": round(dept_avg, 2),
    }


def set_personal_target(user_id: UUID, year: int, month: int,
                         target_score: float, db: Session) -> dict:
    """PB140"""
    existing = db.query(KpiTarget).filter(
        KpiTarget.user_id == user_id,
        KpiTarget.year == year,
        KpiTarget.month == month,
    ).first()

    if existing:
        existing.target_score = target_score
    else:
        db.add(KpiTarget(
            user_id=user_id, year=year,
            month=month, target_score=target_score,
        ))
    db.commit()
    return {"user_id": user_id, "year": year, "month": month, "target_score": target_score}


# ── Manager view KPI ──────────────────────────────────────────

def get_dept_scores(dept_id: UUID, org_id: UUID, year: int, month: int, db: Session) -> list:
    """PB142"""
    members = db.query(User).filter(
        User.dept_id == dept_id,
        User.role == "staff",
        User.is_active == True,
    ).all()

    cfg = _get_config(org_id, db)
    result = []
    for i, m in enumerate(members):
        kpi = calculate_kpi_for_user(m.id, year, month, db)
        result.append({
            "user_id": m.id,
            "full_name": m.full_name,
            "avatar_url": m.avatar_url,
            "total_score": kpi["total_score"],
            "grade": _grade(kpi["total_score"], cfg),
            "rank": 0,
            "breakdown": kpi["breakdown"],
        })

    result.sort(key=lambda x: x["total_score"], reverse=True)
    for i, r in enumerate(result):
        r["rank"] = i + 1

    return result


def get_dept_ranking(dept_id: UUID, org_id: UUID, year: int, month: int, db: Session) -> list:
    """PB143"""
    return get_dept_scores(dept_id, org_id, year, month, db)


def get_grade_distribution(dept_id: UUID, org_id: UUID, year: int, month: int, db: Session) -> dict:
    """PB145"""
    scores = get_dept_scores(dept_id, org_id, year, month, db)
    dist = {"excellent": 0, "good": 0, "pass": 0, "fail": 0, "total": len(scores)}
    for s in scores:
        g = s["grade"]
        if g == "Xuất sắc": dist["excellent"] += 1
        elif g == "Tốt": dist["good"] += 1
        elif g == "Đạt": dist["pass"] += 1
        else: dist["fail"] += 1
    return dist


def get_dept_summary(dept_id: UUID, org_id: UUID, year: int, month: int, db: Session) -> dict:
    """PB141"""
    scores = get_dept_scores(dept_id, org_id, year, month, db)
    avg = sum(s["total_score"] for s in scores) / len(scores) if scores else 0.0
    cfg = _get_config(org_id, db)
    return {
        "year": year, "month": month,
        "member_count": len(scores),
        "average_score": round(avg, 2),
        "summary": {
            "average": round(avg, 2),
            "target": cfg.target_score,
        }
    }


# ── Export Excel ──────────────────────────────────────────────

def export_dept_kpi_excel(dept_id: UUID, org_id: UUID, year: int,
                           month: Optional[int], db: Session) -> bytes:
    """PB148"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    # Sanitize sheet title: replace invalid characters (e.g. '/') and limit length to 31
    safe_title = f"KPI Tháng {month}-{year}" if month else f"KPI Năm {year}"
    for ch in ['\\', '/', '*', '?', ':', '[', ']']:
        safe_title = safe_title.replace(ch, '-')
    safe_title = safe_title[:31]
    ws.title = safe_title

    headers = ["Họ tên", "Tiêu chí", "Điểm", "Trọng số", "Điểm có trọng số", "Tổng", "Xếp loại"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F46E5")

    scores = get_dept_scores(dept_id, org_id, year, month or datetime.now(timezone.utc).month, db)
    for s in scores:
        for bd in s["breakdown"]:
            ws.append([
                s["full_name"], bd["name"], bd["score"],
                bd["weight"], bd["weighted_score"],
                s["total_score"], s["grade"],
            ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def export_company_kpi_excel(org_id: UUID, year: int, month: Optional[int], db: Session) -> bytes:
    """PB146, PB147"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    depts = db.query(Department).filter(
        Department.org_id == org_id,
        Department.is_active == True,
    ).all()

    months = [month] if month else list(range(1, 13))

    for m in months:
        ws = wb.create_sheet(title=f"Tháng {m}")
        ws.append(["Phòng ban", "Họ tên", "Tổng điểm", "Xếp loại"])
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F46E5")

        for dept in depts:
            scores = get_dept_scores(dept.id, org_id, year, m, db)
            for s in scores:
                ws.append([dept.name, s["full_name"], s["total_score"], s["grade"]])

    # Xóa sheet mặc định
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Finalize ──────────────────────────────────────────────────

def finalize_kpi(org_id: UUID, year: int, month: int, user_id: UUID, db: Session) -> dict:
    """PB149"""
    now = datetime.now(timezone.utc)

    rec = db.query(KpiFinalize).filter(
        KpiFinalize.org_id == org_id,
        KpiFinalize.year == year,
        KpiFinalize.month == month,
    ).first()

    if not rec:
        rec = KpiFinalize(org_id=org_id, year=year, month=month)
        db.add(rec)

    rec.is_finalized = True
    rec.finalized_by = user_id
    rec.finalized_at = now

    # Đánh dấu tất cả KpiScore là finalized
    db.query(KpiScore).filter(
        KpiScore.year == year,
        KpiScore.month == month,
    ).update({"is_finalized": True})

    db.commit()

    # PB157: gửi thông báo cho tất cả nhân viên
    _notify_all_staff_kpi_finalized(org_id, year, month, db)

    # PB158: kiểm tra cảnh báo nhân viên dưới ngưỡng
    check_consecutive_low_kpi(org_id, year, month, db)

    return {"year": year, "month": month, "finalized": True}


def unlock_kpi(org_id: UUID, year: int, month: int, reason: str,
               user_id: UUID, db: Session) -> dict:
    """PB150"""
    rec = db.query(KpiFinalize).filter(
        KpiFinalize.org_id == org_id,
        KpiFinalize.year == year,
        KpiFinalize.month == month,
    ).first()

    if not rec:
        raise HTTPException(status_code=404, detail="Không tìm thấy bản ghi chốt KPI")

    rec.is_finalized = False
    rec.unlock_reason = reason
    rec.unlocked_by = user_id
    rec.unlocked_at = datetime.now(timezone.utc)

    db.query(KpiScore).filter(
        KpiScore.year == year,
        KpiScore.month == month,
    ).update({"is_finalized": False})

    db.commit()
    return {"year": year, "month": month, "finalized": False}


def _notify_all_staff_kpi_finalized(org_id: UUID, year: int, month: int, db: Session):
    """PB157"""
    staff_list = db.query(User).filter(
        User.org_id == org_id,
        User.role == "staff",
        User.is_active == True,
    ).all()

    for s in staff_list:
        db.add(Notification(
            user_id=s.id,
            type="kpi_finalized",
            title=f"Kết quả KPI tháng {month}/{year} đã được chốt",
            body=f"Điểm KPI tháng {month}/{year} của bạn đã được chốt. Vào mục KPI để xem kết quả.",
        ))
    db.commit()


def check_consecutive_low_kpi(org_id: UUID, year: int, month: int, db: Session) -> int:
    """PB158: cảnh báo Manager khi nhân viên dưới ngưỡng Đạt 2 tháng liên tiếp"""
    cfg = _get_config(org_id, db)
    warned = 0

    # Tháng trước và 2 tháng trước (kiểm tra 2 tháng liên tiếp trước tháng hiện tại)
    prev1_month = month - 1
    prev1_year = year
    if prev1_month <= 0:
        prev1_month += 12
        prev1_year -= 1

    prev2_month = month - 2
    prev2_year = year
    if prev2_month <= 0:
        prev2_month += 12
        prev2_year -= 1

    staff_list = db.query(User).filter(
        User.org_id == org_id,
        User.role == "staff",
        User.is_active == True,
    ).all()

    for s in staff_list:
        # Điểm 2 tháng trước (prev1 = month-1, prev2 = month-2)
        prev1_scores = db.query(KpiScore).filter(
            KpiScore.user_id == s.id,
            KpiScore.year == prev1_year,
            KpiScore.month == prev1_month,
        ).all()
        prev1_total = sum(sc.weighted_score for sc in prev1_scores)

        prev2_scores = db.query(KpiScore).filter(
            KpiScore.user_id == s.id,
            KpiScore.year == prev2_year,
            KpiScore.month == prev2_month,
        ).all()
        prev2_total = sum(sc.weighted_score for sc in prev2_scores)

        # Cả 2 tháng trước đều dưới ngưỡng Đạt
        if (prev1_total < cfg.threshold_pass and prev1_total > 0 and
                prev2_total < cfg.threshold_pass and prev2_total > 0):

            # Tìm manager phòng ban: ưu tiên dept.manager_id, fallback tìm user role=manager
            manager = None
            if s.dept_id:
                dept = db.query(Department).filter(Department.id == s.dept_id).first()
                if dept and dept.manager_id:
                    manager = db.query(User).filter(User.id == dept.manager_id).first()
                elif dept:
                    manager = db.query(User).filter(
                        User.dept_id == dept.id,
                        User.role == 'manager'
                    ).first()

            if manager:
                existing = db.query(Notification).filter(
                    Notification.user_id == manager.id,
                    Notification.type == "kpi_low_consecutive",
                    Notification.body.like(f"%{s.id}%"),
                ).first()
                if not existing:
                    db.add(Notification(
                        user_id=manager.id,
                        type="kpi_low_consecutive",
                        title=f"Cảnh báo: {s.full_name} dưới ngưỡng Đạt 2 tháng liên tiếp",
                        body=f"Nhân viên {s.full_name} (ID: {s.id}) có điểm KPI dưới {cfg.threshold_pass} điểm 2 tháng liên tiếp. Cần can thiệp hỗ trợ.",
                    ))
                    warned += 1
    db.commit()
    return warned


# ── Appeals ───────────────────────────────────────────────────

def create_appeal(user: User, year: int, month: int, criteria_name: str,
                  current_score: float, proposed_score: float, reason: str,
                  db: Session) -> KpiAppeal:
    """PB151"""
    appeal = KpiAppeal(
        user_id=user.id, year=year, month=month,
        criteria_name=criteria_name, current_score=current_score,
        proposed_score=proposed_score, reason=reason,
    )
    db.add(appeal)
    db.commit()
    db.refresh(appeal)
    return appeal


def respond_appeal(appeal_id: UUID, approved: bool, response: str,
                   adjusted_score: Optional[float], manager: User, db: Session) -> dict:
    """PB152"""
    appeal = db.query(KpiAppeal).filter(KpiAppeal.id == appeal_id).first()
    if not appeal:
        raise HTTPException(status_code=404, detail="Không tìm thấy khiếu nại")

    staff = db.query(User).filter(User.id == appeal.user_id).first()
    if staff.dept_id != manager.dept_id:
        raise HTTPException(status_code=403, detail="Không có quyền phản hồi khiếu nại này")

    appeal.status = "approved" if approved else "rejected"
    appeal.response = response
    appeal.adjusted_score = adjusted_score
    appeal.responded_by = manager.id
    appeal.responded_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(appeal)

    return {
        "id": appeal.id, "status": appeal.status,
        "response": appeal.response, "adjusted_score": appeal.adjusted_score,
    }


# ── Adjustments ───────────────────────────────────────────────

def create_adjustment(manager: User, user_id: UUID, year: int, month: int,
                       criteria_name: str, proposed_score: float,
                       reason: str, db: Session) -> KpiAdjustment:
    """PB153"""
    staff = db.query(User).filter(User.id == user_id).first()
    if not staff or staff.dept_id != manager.dept_id:
        raise HTTPException(status_code=400, detail="Nhân viên không thuộc phòng ban của bạn")

    adj = KpiAdjustment(
        user_id=user_id, requested_by=manager.id,
        year=year, month=month,
        criteria_name=criteria_name,
        proposed_score=proposed_score, reason=reason,
    )
    db.add(adj)
    db.commit()
    db.refresh(adj)
    return adj


def review_adjustment(adj_id: UUID, approved: bool, comment: Optional[str],
                       ceo: User, db: Session) -> dict:
    """PB154"""
    adj = db.query(KpiAdjustment).filter(KpiAdjustment.id == adj_id).first()
    if not adj:
        raise HTTPException(status_code=404, detail="Không tìm thấy yêu cầu điều chỉnh")

    adj.status = "approved" if approved else "rejected"
    adj.comment = comment
    adj.reviewed_by = ceo.id
    adj.reviewed_at = datetime.now(timezone.utc)
    db.commit()

    # PB156: thông báo cho nhân viên
    db.add(Notification(
        user_id=adj.user_id,
        type="kpi_adjustment_result",
        title=f"Kết quả yêu cầu điều chỉnh KPI: {'Đã duyệt' if approved else 'Từ chối'}",
        body=f"Yêu cầu điều chỉnh điểm KPI tiêu chí '{adj.criteria_name}' của bạn đã được {'phê duyệt' if approved else 'từ chối'}. {comment or ''}",
    ))
    db.commit()
    db.refresh(adj)

    requester = db.query(User).filter(User.id == adj.requested_by).first()
    return {
        "id": adj.id,
        "status": adj.status,
        "comment": adj.comment,
        "requester": requester.full_name if requester else "",
        "approver": ceo.full_name,
    }


def get_adjustment_history(org_id: UUID, db: Session) -> list:
    """PB155"""
    adjs = db.query(KpiAdjustment).filter(
        KpiAdjustment.status != "pending"
    ).order_by(KpiAdjustment.reviewed_at.desc()).all()

    result = []
    for a in adjs:
        staff = db.query(User).filter(User.id == a.user_id).first()
        requester = db.query(User).filter(User.id == a.requested_by).first()
        reviewer = db.query(User).filter(User.id == a.reviewed_by).first() if a.reviewed_by else None
        result.append({
            "id": a.id,
            "requester": requester.full_name if requester else "",
            "approver": reviewer.full_name if reviewer else None,
            "staff_name": staff.full_name if staff else "",
            "criteria_name": a.criteria_name,
            "original_score": a.original_score,
            "proposed_score": a.proposed_score,
            "status": a.status,
            "comment": a.comment,
            "created_at": a.created_at,
        })
    return result
