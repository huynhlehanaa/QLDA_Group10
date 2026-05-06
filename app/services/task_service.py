"""
Task service — PB062 đến PB119
"""
import io
from datetime import datetime, timezone, timedelta
from typing import Optional
from uuid import UUID

from fastapi import HTTPException
from sqlalchemy import or_, and_, func
from sqlalchemy.orm import Session

from app.models.task import (
    Task, TaskAssignee, TaskComment, TaskAttachment,
    TaskChecklist, TaskHistory, DeadlineExtensionRequest, Epic,
)
from app.models.user import User
from app.models.organization import Department
from app.schemas.task import (
    TaskCreate, TaskUpdate, TaskFilterParams,
    CommentCreate, ChecklistCreate, ChecklistUpdate,
    ExtensionRequestCreate, ExtensionReview,
)


# ── Helpers ───────────────────────────────────────────────────

def _get_task_or_404(db: Session, task_id: UUID) -> Task:
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Không tìm thấy task")
    return task


def _assert_manager_of_dept(user: User, dept_id: UUID):
    if user.role == "ceo":
        return
    if user.role != "manager" or user.dept_id != dept_id:
        raise HTTPException(status_code=403, detail="Chỉ Manager phòng ban mới thực hiện được")


def _assert_assignee_or_manager(user: User, task: Task, db: Session):
    if user.role in ("ceo", "manager"):
        return
    is_assignee = db.query(TaskAssignee).filter(
        TaskAssignee.task_id == task.id,
        TaskAssignee.user_id == user.id,
    ).first()
    if not is_assignee:
        raise HTTPException(status_code=403, detail="Bạn không có quyền thao tác với task này")


def _log_history(db: Session, task_id: UUID, user_id: UUID,
                 field: str, old: str, new: str, note: str = None):
    """PB092: ghi lịch sử thay đổi"""
    db.add(TaskHistory(
        task_id=task_id, changed_by=user_id,
        field=field, old_value=str(old), new_value=str(new), note=note,
    ))


def _enrich(task: Task, db: Session) -> dict:
    """Thêm thông tin computed vào task response"""
    now = datetime.now(timezone.utc)
    assignees = []
    for ta in task.assignees:
        u = db.query(User).filter(User.id == ta.user_id).first()
        if u:
            assignees.append({"user_id": u.id, "full_name": u.full_name, "avatar_url": u.avatar_url})

    total_cl = len(task.checklists)
    done_cl  = sum(1 for c in task.checklists if c.is_done)

    deadline_aware = task.deadline
    if deadline_aware and deadline_aware.tzinfo is None:
        deadline_aware = deadline_aware.replace(tzinfo=timezone.utc)

    is_overdue = (
        deadline_aware is not None
        and deadline_aware < now
        and task.status not in ("done", "cancelled")
    )
    return {
        "assignees": assignees,
        "checklist_total": total_cl,
        "checklist_done": done_cl,
        "is_overdue": is_overdue,
    }


# ── Epic ──────────────────────────────────────────────────────

def create_epic(name: str, dept_id: UUID, user: User, db: Session) -> Epic:
    """PB077"""
    _assert_manager_of_dept(user, dept_id)
    epic = Epic(name=name, dept_id=dept_id, created_by=user.id)
    db.add(epic)
    db.commit()
    db.refresh(epic)
    return epic


def list_epics(dept_id: UUID, db: Session) -> list:
    """PB077, PB078"""
    epics = db.query(Epic).filter(Epic.dept_id == dept_id).all()
    result = []
    for e in epics:
        tasks = db.query(Task).filter(Task.epic_id == e.id).all()
        total = len(tasks)
        done  = sum(1 for t in tasks if t.status == "done")
        result.append({
            "id": e.id, "name": e.name, "dept_id": e.dept_id,
            "task_count": total, "done_count": done,
            "progress_pct": round(done / total * 100, 1) if total else 0.0,
            "created_at": e.created_at,
        })
    return result


# ── Task CRUD ─────────────────────────────────────────────────

def create_task(data: TaskCreate, user: User, db: Session) -> Task:
    """PB062-PB070, PB073-PB076, PB077"""
    _assert_manager_of_dept(user, user.dept_id)

    # PB064: chỉ giao cho nhân viên trong phòng ban
    if data.assignee_ids:
        for uid in data.assignee_ids:
            assignee = db.query(User).filter(User.id == uid).first()
            if not assignee or assignee.dept_id != user.dept_id:
                raise HTTPException(
                    status_code=400,
                    detail=f"Nhân viên {uid} không thuộc phòng ban của bạn",
                )
            if not assignee.is_active:
                raise HTTPException(status_code=400, detail=f"Nhân viên {assignee.full_name} đang bị vô hiệu hóa")

    # PB103: cảnh báo nhân viên quá tải (> 10 task in_progress)
    overloaded = []
    for uid in data.assignee_ids:
        count = db.query(TaskAssignee).join(Task).filter(
            TaskAssignee.user_id == uid,
            Task.status == "in_progress",
        ).count()
        if count >= 10:
            u = db.query(User).get(uid)
            overloaded.append(u.full_name)
    if overloaded:
        # Vẫn cho tạo nhưng trả cảnh báo qua header (frontend sẽ hiện toast)
        pass  # handled in API layer

    # PB070: validate blocked_by tồn tại
    if data.blocked_by_id:
        blocker = db.query(Task).filter(Task.id == data.blocked_by_id).first()
        if not blocker:
            raise HTTPException(status_code=400, detail="Task phụ thuộc không tồn tại")

    task = Task(
        dept_id       = user.dept_id,
        created_by    = user.id,
        epic_id       = data.epic_id,
        blocked_by_id = data.blocked_by_id,
        title         = data.title,
        description   = data.description,
        priority      = data.priority,
        deadline      = data.deadline,
        is_recurring  = data.is_recurring,
        recur_pattern = data.recur_pattern,
        recur_day     = data.recur_day,
        status        = "todo",
        progress_pct  = 0,
    )
    db.add(task)
    db.flush()

    # PB065: gán nhiều nhân viên
    for uid in data.assignee_ids:
        db.add(TaskAssignee(task_id=task.id, user_id=uid))

    _log_history(db, task.id, user.id, "created", "", task.title)
    db.commit()
    db.refresh(task)
    return task


def get_staff_tasks(staff_id: UUID, requesting_user: User, db: Session, filters: TaskFilterParams) -> dict:
    """
    Manager xem task của một nhân viên cụ thể trong phòng ban.
    Staff chỉ xem được task của chính mình.
    """
    # Kiểm tra quyền
    staff = db.query(User).filter(User.id == staff_id).first()
    if not staff:
        raise HTTPException(status_code=404, detail="Không tìm thấy nhân viên")

    if requesting_user.role == "staff" and requesting_user.id != staff_id:
        raise HTTPException(status_code=403, detail="Bạn chỉ xem được task của chính mình")

    if requesting_user.role == "manager" and staff.dept_id != requesting_user.dept_id:
        raise HTTPException(status_code=403, detail="Nhân viên không thuộc phòng ban của bạn")

    now = datetime.now(timezone.utc)

    def _make_aware(dt):
        if dt and dt.tzinfo is None:
            return dt.replace(tzinfo=timezone.utc)
        return dt

    assigned_ids = [
        ta.task_id for ta in db.query(TaskAssignee).filter(TaskAssignee.user_id == staff_id).all()
    ]
    q = db.query(Task).filter(Task.id.in_(assigned_ids), Task.status != "cancelled")

    if filters.status:
        q = q.filter(Task.status == filters.status)
    if filters.priority:
        q = q.filter(Task.priority == filters.priority)
    if filters.overdue_only:
        q = q.filter(Task.deadline < now, Task.status.notin_(["done", "cancelled"]))

    tasks = q.order_by(Task.deadline.asc().nullslast()).all()
    task_list = [_build_list_item(t, db, now) for t in tasks]

    # Thống kê
    all_tasks = db.query(Task).filter(Task.id.in_(assigned_ids)).all()
    overdue_count = sum(
        1 for t in all_tasks
        if t.deadline and _make_aware(t.deadline) < now
        and t.status not in ("done", "cancelled")
    )

    return {
        "user_id": str(staff.id),
        "full_name": staff.full_name,
        "avatar_url": staff.avatar_url,
        "stats": {
            "total": len(all_tasks),
            "todo": sum(1 for t in all_tasks if t.status == "todo"),
            "in_progress": sum(1 for t in all_tasks if t.status == "in_progress"),
            "done": sum(1 for t in all_tasks if t.status == "done"),
            "overdue": overdue_count,
        },
        "tasks": task_list,
    }


def get_task(task_id: UUID, user: User, db: Session) -> dict:
    """PB083: xem chi tiết task"""
    task = _get_task_or_404(db, task_id)

    # Staff chỉ xem task của mình (PB018)
    if user.role == "staff":
        _assert_assignee_or_manager(user, task, db)

    result = task.__dict__.copy()
    result.update(_enrich(task, db))
    result["checklists"] = [
        {"id": c.id, "content": c.content, "is_done": c.is_done, "position": c.position}
        for c in task.checklists
    ]

    # Comments dạng thread
    top_comments = db.query(TaskComment).filter(
        TaskComment.task_id == task_id,
        TaskComment.parent_id == None,
    ).order_by(TaskComment.created_at).all()

    result["comments"] = [_format_comment(c, db) for c in top_comments]
    result["attachments"] = task.attachments
    result["history"] = _format_history(task.history, db)
    return result


def _format_comment(c: TaskComment, db: Session) -> dict:
    u = db.query(User).filter(User.id == c.user_id).first()
    replies = db.query(TaskComment).filter(TaskComment.parent_id == c.id).all()
    return {
        "id": c.id, "task_id": c.task_id, "user_id": c.user_id,
        "full_name": u.full_name if u else "", "avatar_url": u.avatar_url if u else None,
        "parent_id": c.parent_id, "content": c.content, "created_at": c.created_at,
        "replies": [_format_comment(r, db) for r in replies],
    }


def _format_history(history: list, db: Session) -> list:
    result = []
    for h in sorted(history, key=lambda x: x.created_at, reverse=True):
        u = db.query(User).filter(User.id == h.changed_by).first()
        result.append({
            "id": h.id, "changed_by": h.changed_by,
            "changer_name": u.full_name if u else "",
            "field": h.field, "old_value": h.old_value,
            "new_value": h.new_value, "note": h.note,
            "created_at": h.created_at,
        })
    return result


def list_tasks(user: User, db: Session, filters: TaskFilterParams) -> list:
    """PB079, PB080, PB105, PB108-PB115"""
    q = db.query(Task).filter(Task.status != "cancelled")

    # Phân quyền (PB017, PB018)
    if user.role == "staff":
        assigned_task_ids = db.query(TaskAssignee.task_id).filter(
            TaskAssignee.user_id == user.id
        ).scalar_subquery()
        q = q.filter(Task.id.in_(assigned_task_ids))
    elif user.role == "manager":
        q = q.filter(Task.dept_id == user.dept_id)

    # PB105: tìm theo từ khóa
    if filters.search:
        q = q.filter(Task.title.ilike(f"%{filters.search}%"))

    # PB108: lọc theo status
    if filters.status:
        q = q.filter(Task.status == filters.status)

    # PB110: lọc theo priority
    if filters.priority:
        q = q.filter(Task.priority == filters.priority)

    # PB109: lọc theo assignee (Manager)
    if filters.assignee_id:
        assigned = db.query(TaskAssignee.task_id).filter(
            TaskAssignee.user_id == filters.assignee_id
        ).scalar_subquery()
        q = q.filter(Task.id.in_(assigned))

    # PB111, PB112: lọc deadline
    if filters.deadline_from:
        q = q.filter(Task.deadline >= filters.deadline_from)
    if filters.deadline_to:
        q = q.filter(Task.deadline <= filters.deadline_to)

    # PB115: overdue only
    now = datetime.now(timezone.utc)
    if filters.overdue_only:
        q = q.filter(Task.deadline < now, Task.status.notin_(["done", "cancelled"]))

    # PB113, PB114: sắp xếp
    if filters.sort_by == "priority":
        priority_order = {"high": 0, "medium": 1, "low": 2}
        tasks = q.all()
        tasks.sort(key=lambda t: (priority_order.get(t.priority, 9),
                                   t.deadline or datetime.max.replace(tzinfo=timezone.utc)))
        return [_build_list_item(t, db, now) for t in tasks]

    if filters.sort_dir == "desc":
        q = q.order_by(Task.deadline.desc().nullslast())
    else:
        q = q.order_by(Task.deadline.asc().nullslast())

    tasks = q.all()
    return [_build_list_item(t, db, now) for t in tasks]


def _build_list_item(task: Task, db: Session, now: datetime) -> dict:
    enriched = _enrich(task, db)
    deadline_aware = task.deadline
    if deadline_aware and deadline_aware.tzinfo is None:
        deadline_aware = deadline_aware.replace(tzinfo=timezone.utc)
    return {
        "id": task.id, "title": task.title, "status": task.status,
        "priority": task.priority, "progress_pct": task.progress_pct,
        "deadline": task.deadline, "created_at": task.created_at,
        "epic_id": task.epic_id,
        **enriched,
    }


def get_kanban(user: User, db: Session, filters: TaskFilterParams) -> dict:
    """PB081, PB082: Kanban 3 cột"""
    all_tasks = list_tasks(user, db, filters)
    columns = {"todo": [], "in_progress": [], "done": []}
    for t in all_tasks:
        s = t["status"]
        if s in columns:
            columns[s].append(t)
    return {
        status: {"status": status, "count": len(items), "tasks": items}
        for status, items in columns.items()
    }


def update_task(task_id: UUID, data: TaskUpdate, user: User, db: Session) -> Task:
    """PB094, PB095, PB096"""
    task = _get_task_or_404(db, task_id)
    _assert_manager_of_dept(user, task.dept_id)

    if data.title is not None:
        _log_history(db, task_id, user.id, "title", task.title, data.title)
        task.title = data.title

    if data.description is not None:
        task.description = data.description

    if data.priority is not None:
        _log_history(db, task_id, user.id, "priority", task.priority, data.priority)
        task.priority = data.priority

    if data.deadline is not None:                    # PB095
        _log_history(db, task_id, user.id, "deadline",
                     str(task.deadline), str(data.deadline),
                     note=data.deadline_change_reason)
        task.deadline = data.deadline

    if data.assignee_ids is not None:                # PB096
        db.query(TaskAssignee).filter(TaskAssignee.task_id == task_id).delete()
        for uid in data.assignee_ids:
            db.add(TaskAssignee(task_id=task_id, user_id=uid))
        _log_history(db, task_id, user.id, "assignees",
                     "", ", ".join(str(x) for x in data.assignee_ids))

    db.commit()
    db.refresh(task)
    return task


def update_status(task_id: UUID, status: str, progress_pct: Optional[int],
                  user: User, db: Session) -> Task:
    """PB084, PB085, PB086"""
    task = _get_task_or_404(db, task_id)
    _assert_assignee_or_manager(user, task, db)

    old_status = task.status
    now = datetime.now(timezone.utc)

    # PB086: done → ghi timestamp
    if status == "done":
        task.completed_at = now
        task.progress_pct = 100

    # PB085: progress 100 → tự chuyển done
    if progress_pct is not None:
        task.progress_pct = progress_pct
        if progress_pct == 100 and status != "done":
            status = "done"
            task.completed_at = now

    if status != old_status:
        _log_history(db, task_id, user.id, "status", old_status, status)
        task.status = status

    db.commit()
    db.refresh(task)
    return task


def cancel_task(task_id: UUID, reason: str, user: User, db: Session) -> Task:
    """PB097"""
    task = _get_task_or_404(db, task_id)
    _assert_manager_of_dept(user, task.dept_id)

    _log_history(db, task_id, user.id, "status", task.status, "cancelled", note=reason)
    task.status = "cancelled"
    task.cancel_reason = reason
    task.cancelled_at = datetime.now(timezone.utc)

    db.commit()
    db.refresh(task)
    return task


def stop_recurring(task_id: UUID, user: User, db: Session) -> Task:
    """PB076"""
    task = _get_task_or_404(db, task_id)
    _assert_manager_of_dept(user, task.dept_id)
    task.is_recurring = False
    task.recur_pattern = None
    db.commit()
    db.refresh(task)
    return task


# ── Comment ───────────────────────────────────────────────────

def add_comment(task_id: UUID, data: CommentCreate, user: User, db: Session) -> TaskComment:
    """PB087, PB088, PB091"""
    task = _get_task_or_404(db, task_id)
    _assert_assignee_or_manager(user, task, db)

    # PB087: lần comment đầu tiên → tự chuyển sang in_progress
    auto_transitioned = False
    if task.status == "todo":
        existing = db.query(TaskComment).filter(TaskComment.task_id == task_id).count()
        if existing == 0:
            _log_history(db, task_id, user.id, "status", "todo", "in_progress",
                         note="Tự động chuyển khi thêm ghi chú đầu tiên")
            task.status = "in_progress"
            auto_transitioned = True

    comment = TaskComment(
        task_id=task_id, user_id=user.id,
        content=data.content, parent_id=data.parent_id,
    )
    db.add(comment)
    db.commit()
    db.refresh(comment)
    return comment


def list_comments(task_id: UUID, db: Session) -> list:
    """PB090"""
    comments = db.query(TaskComment).filter(
        TaskComment.task_id == task_id,
        TaskComment.parent_id == None,
    ).order_by(TaskComment.created_at).all()
    return [_format_comment(c, db) for c in comments]


# ── Attachment ────────────────────────────────────────────────

def add_attachment(task_id: UUID, file_url: str, file_name: str,
                   file_size: int, user: User, db: Session) -> TaskAttachment:
    """PB069, PB089"""
    task = _get_task_or_404(db, task_id)
    _assert_assignee_or_manager(user, task, db)

    # PB069: tối đa 5 file/task
    count = db.query(TaskAttachment).filter(TaskAttachment.task_id == task_id).count()
    if count >= 5:
        raise HTTPException(status_code=400, detail="Task chỉ được đính kèm tối đa 5 file")

    # PB069: tối đa 10MB
    if file_size and file_size > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File không được vượt quá 10MB")

    att = TaskAttachment(
        task_id=task_id, uploaded_by=user.id,
        file_url=file_url, file_name=file_name, file_size=file_size,
    )
    db.add(att)
    db.commit()
    db.refresh(att)
    return att


def delete_attachment(att_id: UUID, user: User, db: Session):
    att = db.query(TaskAttachment).filter(TaskAttachment.id == att_id).first()
    if not att:
        raise HTTPException(status_code=404, detail="Không tìm thấy file")
    if att.uploaded_by != user.id and user.role not in ("manager", "ceo"):
        raise HTTPException(status_code=403, detail="Không có quyền xóa file này")
    db.delete(att)
    db.commit()


# ── Checklist ─────────────────────────────────────────────────

def add_checklist(task_id: UUID, data: ChecklistCreate, user: User, db: Session) -> TaskChecklist:
    """PB071"""
    task = _get_task_or_404(db, task_id)
    _assert_assignee_or_manager(user, task, db)
    item = TaskChecklist(task_id=task_id, content=data.content, position=data.position)
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


def update_checklist(item_id: UUID, data: ChecklistUpdate, user: User, db: Session) -> TaskChecklist:
    """PB071, PB072"""
    item = db.query(TaskChecklist).filter(TaskChecklist.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Không tìm thấy checklist item")
    task = _get_task_or_404(db, item.task_id)
    _assert_assignee_or_manager(user, task, db)

    if data.is_done is not None:
        item.is_done = data.is_done
    if data.content is not None:
        item.content = data.content

    # PB072: tự cập nhật % progress theo checklist
    checklists = db.query(TaskChecklist).filter(TaskChecklist.task_id == item.task_id).all()
    if checklists:
        # db.refresh sau khi item.is_done đã được set ở trên
        done_count = sum(1 for c in checklists if c.is_done)
        task.progress_pct = round(done_count / len(checklists) * 100)

    db.commit()
    db.refresh(item)
    return item


def delete_checklist(item_id: UUID, user: User, db: Session):
    item = db.query(TaskChecklist).filter(TaskChecklist.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404)
    db.delete(item)
    db.commit()


# ── Deadline extension ────────────────────────────────────────

def request_extension(task_id: UUID, data: ExtensionRequestCreate,
                      user: User, db: Session) -> DeadlineExtensionRequest:
    """PB099"""
    task = _get_task_or_404(db, task_id)
    _assert_assignee_or_manager(user, task, db)
    req = DeadlineExtensionRequest(
        task_id=task_id, requested_by=user.id,
        proposed_deadline=data.proposed_deadline, reason=data.reason,
    )
    db.add(req)
    db.commit()
    db.refresh(req)
    return {
        "id": str(req.id),
        "task_id": str(req.task_id),
        "proposed_deadline": str(req.proposed_deadline),
        "reason": req.reason,
        "status": req.status,
        "created_at": str(req.created_at),
    }


def review_extension(req_id: UUID, data: ExtensionReview, user: User, db: Session):
    """PB100"""
    req = db.query(DeadlineExtensionRequest).filter(
        DeadlineExtensionRequest.id == req_id
    ).first()
    if not req:
        raise HTTPException(status_code=404, detail="Không tìm thấy yêu cầu gia hạn")

    task = _get_task_or_404(db, req.task_id)
    _assert_manager_of_dept(user, task.dept_id)

    req.status = "approved" if data.approved else "rejected"
    req.reviewed_by = user.id
    req.review_note = data.note
    req.reviewed_at = datetime.now(timezone.utc)

    if data.approved:
        old_deadline = task.deadline
        task.deadline = req.proposed_deadline
        _log_history(db, task.id, user.id, "deadline",
                     str(old_deadline), str(req.proposed_deadline),
                     note=f"Phê duyệt gia hạn: {data.note}")

    db.commit()
    db.refresh(req)
    return {
        "id": str(req.id),
        "task_id": str(req.task_id),
        "status": req.status,
        "review_note": req.review_note,
        "reviewed_at": str(req.reviewed_at),
    }


# ── Workload ──────────────────────────────────────────────────

def get_workload(dept_id: UUID, user: User, db: Session) -> list:
    """PB101, PB102"""
    _assert_manager_of_dept(user, dept_id)
    now = datetime.now(timezone.utc)

    staff_members = db.query(User).filter(
        User.dept_id == dept_id,
        User.role == "staff",
        User.is_active == True,
    ).all()

    result = []
    for s in staff_members:
        assigned_ids = [ta.task_id for ta in db.query(TaskAssignee).filter(
            TaskAssignee.user_id == s.id
        ).all()]
        tasks = db.query(Task).filter(
            Task.id.in_(assigned_ids),
            Task.status != "cancelled",
        ).all()

        todo_count = sum(1 for t in tasks if t.status == "todo")
        ip_count   = sum(1 for t in tasks if t.status == "in_progress")
        done_count = sum(1 for t in tasks if t.status == "done")

        def _make_aware(dt):
            if dt and dt.tzinfo is None:
                return dt.replace(tzinfo=timezone.utc)
            return dt

        overdue = sum(1 for t in tasks if t.deadline and _make_aware(t.deadline) < now
                      and t.status not in ("done", "cancelled"))

        result.append({
            "user_id": s.id, "full_name": s.full_name, "avatar_url": s.avatar_url,
            "todo_count": todo_count, "in_progress_count": ip_count,
            "done_count": done_count, "overdue_count": overdue,
            "total": len(tasks),
        })

    return sorted(result, key=lambda x: x["total"], reverse=True)


# ── Stats ─────────────────────────────────────────────────────

def get_task_stats(user: User, db: Session,
                   from_date: Optional[datetime] = None,
                   to_date: Optional[datetime] = None) -> dict:
    """PB117, PB118"""
    now = datetime.now(timezone.utc)

    def _make_aware(dt):
        if dt and dt.tzinfo is None:
            return dt.replace(tzinfo=timezone.utc)
        return dt
    q = db.query(Task)

    if user.role == "staff":
        assigned = db.query(TaskAssignee.task_id).filter(TaskAssignee.user_id == user.id).scalar_subquery()
        q = q.filter(Task.id.in_(assigned))
    elif user.role == "manager":
        q = q.filter(Task.dept_id == user.dept_id)

    if from_date:
        q = q.filter(Task.created_at >= from_date)
    if to_date:
        q = q.filter(Task.created_at <= to_date)

    tasks = q.all()
    done_on_time = sum(1 for t in tasks if t.status == "done"
                       and t.completed_at and t.deadline
                       and t.completed_at <= t.deadline)
    done_late    = sum(1 for t in tasks if t.status == "done"
                       and t.completed_at and t.deadline
                       and t.completed_at > t.deadline)

    overdue     =  sum(1 for t in tasks if t.deadline and _make_aware(t.deadline) < now
                       and t.status not in ("done", "cancelled"))

    return {
        "total": len(tasks),
        "done_on_time": done_on_time,
        "done_late": done_late,
        "in_progress": sum(1 for t in tasks if t.status == "in_progress"),
        "overdue": overdue,
        "cancelled": sum(1 for t in tasks if t.status == "cancelled"),
    }


# ── Export Excel ──────────────────────────────────────────────

def export_tasks_excel(user: User, db: Session, filters: TaskFilterParams) -> bytes:
    """PB119"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    tasks = list_tasks(user, db, filters)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách Task"

    headers = ["ID", "Tiêu đề", "Người thực hiện", "Trạng thái",
               "Độ ưu tiên", "Deadline", "% Hoàn thành", "Quá hạn"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5")
    for col, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 20

    status_map = {"todo": "Chưa làm", "in_progress": "Đang làm",
                  "done": "Hoàn thành", "cancelled": "Đã hủy"}
    priority_map = {"low": "Thấp", "medium": "Trung bình", "high": "Cao"}

    for t in tasks:
        assignee_names = ", ".join(a["full_name"] for a in t.get("assignees", []))
        deadline_str = t["deadline"].strftime("%d/%m/%Y %H:%M") if t["deadline"] else ""
        ws.append([
            str(t["id"])[:8] + "...",
            t["title"],
            assignee_names,
            status_map.get(t["status"], t["status"]),
            priority_map.get(t["priority"], t["priority"]),
            deadline_str,
            f'{t["progress_pct"]}%',
            "Có" if t["is_overdue"] else "Không",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
